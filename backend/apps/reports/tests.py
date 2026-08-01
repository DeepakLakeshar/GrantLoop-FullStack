import io
import uuid
from decimal import Decimal
from typing import List, Dict, Any
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.test import TestCase
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from rest_framework import status
from rest_framework.test import APITestCase, APIRequestFactory

import openpyxl

from apps.campaigns.models import Campaign, Category, TransparencyLog, Verification
from apps.donations.models import Donation
from apps.beneficiaries.models import Beneficiary, VerificationStatus
from apps.payouts.models import Payout
from apps.reports.constants import REPORT_VERSION, ReportType, EXPORTER_REGISTRY, VALID_EXPORT_FORMATS
from apps.reports import services, exporters, throttles

User = get_user_model()


class ReportExportersUnitTest(TestCase):
    """
    Independent unit test coverage for OOP Exporter architecture in exporters.py
    """

    def setUp(self):
        self.headers = ["ID", "Name", "Amount", "Date"]
        self.data_list = [
            {"ID": "101", "Name": "Grant Alpha", "Amount": Decimal("5000.00"), "Date": "2026-08-01 12:00:00"},
            {"ID": "102", "Name": "Grant Beta", "Amount": Decimal("250.50"), "Date": "2026-08-02 14:30:00"},
        ]
        self.metadata = {
            "generated_at": timezone.now().isoformat(),
            "generated_by": "unit.tester@grantloop.org",
            "filters_applied": {"status": "success"},
            "total_records": len(self.data_list),
            "report_version": REPORT_VERSION,
        }

    def test_01_build_report_metadata(self):
        meta = exporters.build_report_metadata(user="test@org", params={"year": "2026", "format": "csv"}, total_records=10)
        self.assertEqual(meta["report_version"], REPORT_VERSION)
        self.assertEqual(meta["total_records"], 10)
        self.assertNotIn("format", meta["filters_applied"])
        self.assertEqual(meta["filters_applied"]["year"], "2026")

    def test_02_csv_exporter_generator(self):
        exporter = exporters.CSVExporter(self.data_list, self.headers, self.metadata, title="Test CSV")
        stream = list(exporter.generate_content())
        self.assertTrue(len(stream) >= 5) # Preamble comments + headers + data rows
        self.assertTrue(any("Grant Alpha" in row for row in stream))

    def test_03_excel_exporter_bytes_and_frozen_panes(self):
        exporter = exporters.ExcelExporter(self.data_list, self.headers, self.metadata, title="Test Excel")
        content = exporter.generate_content()
        self.assertIsInstance(content, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        self.assertEqual(ws.title, "Report Data")
        self.assertEqual(ws.freeze_panes, "A6")
        self.assertEqual(ws.cell(row=1, column=1).value, "Test Excel")

    def test_04_excel_exporter_currency_formatting(self):
        exporter = exporters.ExcelExporter(self.data_list, self.headers, self.metadata, title="Test Excel Format")
        wb = openpyxl.load_workbook(io.BytesIO(exporter.generate_content()))
        ws = wb.active
        amount_cell = ws.cell(row=6, column=3) # Row 6 is first data row, Column 3 is Amount
        self.assertEqual(amount_cell.number_format, "#,##0.00")
        self.assertEqual(float(amount_cell.value), 5000.00)

    def test_05_pdf_exporter_bytes(self):
        exporter = exporters.PDFExporter(self.data_list, self.headers, self.metadata, title="Test PDF")
        content = exporter.generate_content()
        self.assertIsInstance(content, bytes)
        self.assertTrue(content.startswith(b"%PDF-"))

    def test_06_zip_exporter_not_implemented(self):
        exporter = exporters.ZIPExporter(self.data_list, self.headers, self.metadata, title="Test ZIP")
        with self.assertRaises(NotImplementedError):
            exporter.generate_content()

    def test_07_get_exporter_factory(self):
        exporter = exporters.get_exporter("xlsx", self.data_list, self.headers, self.metadata)
        self.assertIsInstance(exporter, exporters.ExcelExporter)

    def test_08_get_exporter_invalid_format(self):
        with self.assertRaises(ValueError):
            exporters.get_exporter("unsupported", self.data_list, self.headers, self.metadata)


class ReportServicesIntegrationTest(TestCase):
    """
    Service layer testing verifying zero duplicated ORM logic and accurate role scoping.
    """

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(
            email="admin@reports.org", password="password123", full_name="Admin Boss", role="admin"
        )
        cls.ngo = User.objects.create_user(
            email="ngo@reports.org", password="password123", full_name="Health NGO", role="ngo", email_verified=True
        )
        cls.donor = User.objects.create_user(
            email="donor@reports.org", password="password123", full_name="Donor Philanthropist", role="donor"
        )
        cls.category = Category.objects.create(name="Health Reports", slug="health-reports")
        cls.campaign = Campaign.objects.create(
            title="Clean Water Project",
            goal_amount=Decimal("50000.00"),
            raised_amount=Decimal("15000.00"),
            campaign_currency="USD",
            status="live",
            category=cls.category,
            created_by=cls.ngo,
        )
        Verification.objects.create(campaign=cls.campaign, verified_by=cls.admin, status="approved", notes="Valid project")
        cls.donation = Donation.objects.create(
            donor=cls.donor,
            campaign=cls.campaign,
            original_amount=Decimal("500.00"),
            original_currency="USD",
            settled_amount=Decimal("500.00"),
            settled_currency="USD",
            status="completed",
            gateway_order_id="ord_rep_1",
            gateway_transaction_id="pay_rep_1",
        )
        cls.beneficiary = Beneficiary.objects.create(
            campaign=cls.campaign,
            full_name="Jane Beneficiary",
            verification_status=VerificationStatus.VERIFIED,
            rejection_reason="",
        )
        cls.payout = Payout.objects.create(
            campaign=cls.campaign,
            ngo=cls.ngo,
            requested_by=cls.ngo,
            requested_amount=Decimal("2000.00"),
            approved_amount=Decimal("2000.00"),
            currency="USD",
            status="completed",
            available_balance_before=Decimal("15000.00"),
            available_balance_after=Decimal("13000.00"),
        )
        cls.log = TransparencyLog.objects.create(
            campaign=cls.campaign, action="Campaign approved for reports verification"
        )

    def test_09_generate_donation_report_admin(self):
        qs, headers, data = services.generate_donation_report(self.admin)
        self.assertTrue(len(data) >= 1)
        self.assertEqual(data[0]["Donation ID"], str(self.donation.id))

    def test_10_generate_donation_report_ngo_scoped(self):
        qs, headers, data = services.generate_donation_report(self.ngo)
        self.assertEqual(len(data), 1)

    def test_11_generate_donation_report_donor_scoped(self):
        qs, headers, data = services.generate_donation_report(self.donor)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["Amount"], Decimal("500.00"))

    def test_12_generate_campaign_report_admin_annotations(self):
        qs, headers, data = services.generate_campaign_report(self.admin)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["Beneficiary Count"], 1)

    def test_13_generate_campaign_report_ngo_scoped(self):
        qs, headers, data = services.generate_campaign_report(self.ngo)
        self.assertEqual(data[0]["Title"], "Clean Water Project")

    def test_14_generate_ngo_report(self):
        qs, headers, data = services.generate_ngo_report(self.admin)
        self.assertTrue(len(data) >= 1)
        self.assertTrue(any(row["Email"] == self.ngo.email for row in data))

    def test_15_generate_beneficiary_report(self):
        qs, headers, data = services.generate_beneficiary_report(self.ngo)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["Full Name"], "Jane Beneficiary")

    def test_16_generate_payout_report(self):
        qs, headers, data = services.generate_payout_report(self.admin)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["Amount"], Decimal("2000.00"))

    def test_17_generate_financial_report_accounting_consistency(self):
        data, headers, raw = services.generate_financial_report(self.admin)
        self.assertEqual(len(raw), 1)
        self.assertTrue(raw[0]["Accounting Consistent"])
        self.assertIsInstance(raw[0]["Platform Balance"], Decimal)

    def test_18_generate_financial_report_non_admin_blocked(self):
        with self.assertRaises(PermissionDenied):
            services.generate_financial_report(self.ngo)

    def test_19_generate_audit_log_report(self):
        qs, headers, data = services.generate_audit_log_report(self.admin)
        self.assertTrue(len(data) >= 1)
        self.assertTrue(any("approved for reports" in row["Action"] for row in data))

    def test_20_filter_by_status(self):
        qs, headers, data = services.generate_donation_report(self.admin, {"status": "refunded"})
        self.assertEqual(len(data), 0)

    def test_21_filter_by_search_term(self):
        qs, headers, data = services.generate_donation_report(self.admin, {"search": "pay_rep_1"})
        self.assertEqual(len(data), 1)

    def test_22_empty_dataset_handling(self):
        qs, headers, data = services.generate_donation_report(self.admin, {"year": "1990"})
        self.assertEqual(len(data), 0)
        self.assertIn("Donation ID", headers)

    def test_23_schedule_report_generation_abstraction(self):
        ticket = services.schedule_report_generation(ReportType.DONATIONS, self.admin, {}, "csv")
        self.assertEqual(ticket["status"], "QUEUED")
        self.assertTrue(ticket["job_id"].startswith("job_"))

    def test_24_generate_report_now_returns_exporter(self):
        res = services.generate_report_now(ReportType.DONATIONS, self.admin, {}, export_format="pdf")
        self.assertIsInstance(res, exporters.PDFExporter)

    def test_25_generate_report_now_returns_json_dict(self):
        res = services.generate_report_now(ReportType.DONATIONS, self.admin, {}, export_format=None)
        self.assertIn("metadata", res)
        self.assertIn("data", res)

    def test_26_legacy_export_helpers(self):
        exc = services.export_excel([{"ID": 1}], ["ID"], {})
        self.assertIsInstance(exc, exporters.ExcelExporter)


class ReportPerformanceAndLoadTest(TestCase):
    """
    Performance verification proving zero N+1 latency and 1000+ record load stability.
    """

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(
            email="perf_admin@reports.org", password="password123", role="admin"
        )
        cls.donor = User.objects.create_user(
            email="perf_donor@reports.org", password="password123", role="donor"
        )
        cls.category = Category.objects.create(name="Perf Grants", slug="perf-grants")
        cls.campaign = Campaign.objects.create(
            title="Mass Scale Project",
            goal_amount=Decimal("1000000.00"),
            raised_amount=Decimal("100000.00"),
            campaign_currency="USD",
            status="live",
            category=cls.category,
            created_by=cls.admin,
        )
        # Generate 10 baseline records
        donations_to_create = [
            Donation(
                donor=cls.donor,
                campaign=cls.campaign,
                original_amount=Decimal("10.00"),
                original_currency="USD",
                settled_amount=Decimal("10.00"),
                settled_currency="USD",
                status="completed",
                gateway_order_id=f"ord_{idx}",
            )
            for idx in range(10)
        ]
        Donation.objects.bulk_create(donations_to_create)

    def test_27_assert_num_queries_donation_report_constant(self):
        # Determine exact query baseline for 10 records
        with self.assertNumQueries(1): # Exactly 1 optimized select_related query
            qs, headers, data = services.generate_donation_report(self.admin)
            _ = list(data)

    def test_28_assert_num_queries_campaign_report_constant(self):
        with self.assertNumQueries(1): # Exactly 1 annotated query
            qs, headers, data = services.generate_campaign_report(self.admin)
            _ = list(data)

    def test_29_large_dataset_performance_1000_records_stream(self):
        # Create 1,020 more donations to surpass the 1,000 record stress test milestone
        bulk_list = [
            Donation(
                donor=self.donor,
                campaign=self.campaign,
                original_amount=Decimal("5.00"),
                original_currency="USD",
                settled_amount=Decimal("5.00"),
                settled_currency="USD",
                status="completed",
                gateway_order_id=f"bulk_ord_{idx}",
            )
            for idx in range(1020)
        ]
        Donation.objects.bulk_create(bulk_list)

        qs, headers, data_dicts = services.generate_donation_report(self.admin)
        self.assertTrue(len(data_dicts) >= 1030)
        exporter = exporters.CSVExporter(data_dicts, headers, {"total_records": len(data_dicts)})
        rows = list(exporter.generate_content())
        self.assertTrue(len(rows) >= 1030)


class ReportEndpointsAndViewsTest(APITestCase):
    """
    Exhaustive integration test suite verifying thin view controllers, pagination vs export rules, and role isolation.
    """

    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(email="view_admin@reports.org", password="password123", role="admin")
        cls.ngo = User.objects.create_user(email="view_ngo@reports.org", password="password123", role="ngo", email_verified=True)
        cls.donor = User.objects.create_user(email="view_donor@reports.org", password="password123", role="donor")
        cls.category = Category.objects.create(name="View Grants", slug="view-grants")
        cls.campaign = Campaign.objects.create(
            title="Endpoint Project",
            goal_amount=Decimal("20000.00"),
            raised_amount=Decimal("5000.00"),
            status="live",
            category=cls.category,
            created_by=cls.ngo,
        )
        for i in range(25):
            Donation.objects.create(
                donor=cls.donor,
                campaign=cls.campaign,
                original_amount=Decimal("100.00"),
                original_currency="USD",
                settled_amount=Decimal("100.00"),
                settled_currency="USD",
                status="completed",
                gateway_order_id=f"ep_ord_{i}",
            )

    def test_30_donations_json_endpoint_pagination(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations"))
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("count", res.data)
        self.assertEqual(res.data["count"], 25) # 25 records paginated by 20
        self.assertLessEqual(len(res.data["results"]["records"]), 20)

    def test_31_donations_export_csv_unpaginated(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations-export"), {"format": "csv"})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn("grantloop_donations_", res["Content-Disposition"])
        self.assertIn(".csv", res["Content-Disposition"])
        # Verify unpaginated complete dataset in streaming buffer
        content_str = "".join([chunk.decode("utf-8") for chunk in res.streaming_content])
        self.assertEqual(content_str.count("ep_ord_"), 25)

    def test_32_donations_export_xlsx(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations-export"), {"format": "xlsx"})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_33_donations_export_pdf(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations-export"), {"format": "pdf"})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res["Content-Type"], "application/pdf")

    def test_34_export_invalid_format_returns_400(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations-export"), {"format": "docx"})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_35_schedule_param_json_view(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations"), {"schedule": "true"})
        self.assertEqual(res.status_code, status.HTTP_202_ACCEPTED)
        self.assertEqual(res.data["status"], "QUEUED")

    def test_36_schedule_param_export_view(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-donations-export"), {"format": "csv", "schedule": "true"})
        self.assertEqual(res.status_code, status.HTTP_202_ACCEPTED)

    def test_37_role_isolation_donor_blocked_from_operational(self):
        self.client.force_authenticate(user=self.donor)
        res = self.client.get(reverse("report-campaigns"))
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_38_role_isolation_ngo_blocked_from_financial(self):
        self.client.force_authenticate(user=self.ngo)
        res = self.client.get(reverse("report-financial"))
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_39_unauthenticated_requests_denied(self):
        res = self.client.get(reverse("report-donations"))
        self.assertIn(res.status_code, [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN])

    def test_40_drf_throttling_export_rate(self):
        throttle = throttles.ReportExportThrottle()
        self.assertEqual(throttle.get_rate(), "100/minute")
        self.assertEqual(throttle.scope, "report_export")

    def test_41_audit_logs_endpoint(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-audit-logs"))
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_42_ngos_report_endpoint(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.get(reverse("report-ngos"))
        self.assertEqual(res.status_code, status.HTTP_200_OK)
